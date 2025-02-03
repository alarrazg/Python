import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

# Datos de viabilidad de los juegos
juegos = [
    "Assassin’s Creed Valhalla", "Cyberpunk 2077", "Far Cry 6", "Watch Dogs Legion", 
    "Battlefield V", "Fortnite", "League of Legends", "Valorant", "Rocket League", 
    "Minecraft", "Shadow of the Tomb Raider", "Forza Horizon 5", "Call of Duty: Warzone", 
    "Doom Eternal", "Watch Dogs Legion", "Apex Legends", "Call of Duty: Modern Warfare II", 
    "Horizon Zero Dawn", "Battlefield 2042"
]

# Combinaciones de tarjetas gráficas
combinaciones = [
    "AMD Radeon RX 6600", "NVIDIA GeForce GTX 1660 Super", "AMD Radeon RX 6650 XT"
]

# Puntuaciones de viabilidad (1-5) según la relación de cada combinación de hardware y juego
viabilidad = {
    "Assassin’s Creed Valhalla": [4, 3, 4],
    "Cyberpunk 2077": [2, 1, 3],
    "Far Cry 6": [4, 3, 4],
    "Watch Dogs Legion": [4, 3, 4],
    "Battlefield V": [5, 4, 5],
    "Fortnite": [5, 4, 5],
    "League of Legends": [5, 5, 5],
    "Valorant": [5, 4, 5],
    "Rocket League": [5, 4, 5],
    "Minecraft": [5, 4, 5],
    "Shadow of the Tomb Raider": [4, 3, 4],
    "Forza Horizon 5": [4, 3, 4],
    "Call of Duty: Warzone": [3, 3, 3],
    "Doom Eternal": [5, 4, 5],
    "Watch Dogs Legion": [4, 3, 4],
    "Apex Legends": [4, 3, 4],
    "Call of Duty: Modern Warfare II": [4, 3, 4],
    "Horizon Zero Dawn": [5, 3, 4],
    "Battlefield 2042": [3, 3, 3]
}

# Crear el DataFrame con las combinaciones y la viabilidad de los juegos
df_viabilidad = pd.DataFrame(viabilidad, index=combinaciones)

# Guardar el archivo Excel
file_path_viabilidad = "P:/Mis_Cosas/viabilidad_juegos_combinaciones.xlsx"
df_viabilidad.to_excel(file_path_viabilidad)

# Abrir el archivo Excel para aplicar formato condicional
wb = openpyxl.load_workbook(file_path_viabilidad)
ws = wb.active

# Crear un formato de escala de colores (rojo, amarillo, verde) con puntuaciones 1-5
color_map = {
    1: PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # Rojo
    2: PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid"),  # Naranja
    3: PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),  # Amarillo
    4: PatternFill(start_color="99FF00", end_color="99FF00", fill_type="solid"),  # Verde claro
    5: PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")   # Verde
}

# Aplicar la escala de colores según las puntuaciones
for row in ws.iter_rows(min_row=2, min_col=2, max_row=len(combinaciones) + 1, max_col=len(juegos) + 1):
    for cell in row:
        score = cell.value
        fill_color = color_map.get(score, PatternFill())  # Si la puntuación no está definida, no aplicar color
        cell.fill = fill_color

# Guardar el archivo con el formato de colores aplicado
wb.save(file_path_viabilidad)

# Proporcionar el enlace para descarga
file_path_viabilidad
